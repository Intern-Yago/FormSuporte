# views.py
from datetime import date, datetime, timedelta
import json
import os
import unicodedata
import requests

from dateutil.relativedelta import relativedelta
from email.utils import parsedate_to_datetime

from django.conf import settings
from django.contrib.admin.models import CHANGE, ADDITION
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import IntegrityError
from django.db.models import Q
from django.http import JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from ocorrencia_erro.models import Device
from painel.decorators import require_system_access
from situacao_veiculo.audit import admin_log

from .models import Cliente,SerialSearchLog
from .services.odoo_sync import sync_odoo_to_clientes


# =========================================================
# Helpers: serial principal OU secundário
# =========================================================
# Helpers: serial principal OU secundário OU e-mail
def _qs_by_any_id(query: str):
    query = (query or "").strip()
    if not query:
        return Cliente.objects.none()
    return Cliente.objects.filter(
        Q(serial__iexact=query) | 
        Q(serial_sec__iexact=query) |
        Q(email__iexact=query)
    )

def _first_by_any_id(query: str):
    return _qs_by_any_id(query).order_by("-updated_at", "-id").first()


def check_user_full_permission(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True

    profile = getattr(user, 'profile', None)
    if not profile:
        return False

    if profile.role in ['dono', 'diretor', 'ti']:
        return True
    if profile.role == 'gestor' and profile.setor == 'suporte':
        return True
    return False


def buscar_serial(request):
    context = {'is_gestor_suporte_ou_master': check_user_full_permission(request.user)}
    if request.method == 'POST':
        serial_input = (request.POST.get('serial') or '').strip()
        user = request.user
        context['serial_digitado'] = serial_input

        clientes = _qs_by_any_id(serial_input)

        # =======================
        # NÃO ENCONTROU NO BANCO
        # =======================
        if not clientes.exists():
            try:
                cookie_value = 'eyJ1c2VyX2lkIjoiZWFhdGFkbWluIn0.aV_2TA.URbvR1qfUJFd6H56IRWZc_hSSp0'
                headers = {
                    'Accept': '*/*',
                    'Accept-Language': 'pt-PT,pt;q=0.9,en-US;q=0.8,en;q=0.7,es;q=0.6',
                    'Connection': 'keep-alive',
                    'Content-Type': 'application/json',
                    'Origin': 'http://20.83.150.13:8088',
                    'Referer': 'http://20.83.150.13:8088/dashboard',
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36',
                }
                resp = requests.post(
                    'http://20.83.150.13:8088/search_codes',
                    json={'sn': serial_input},
                    headers=headers,
                    cookies={'session': cookie_value},
                    timeout=10,
                )

                try:
                    resp_json = resp.json()
                except ValueError:
                    resp_json = {'status_code': resp.status_code, 'text': resp.text}

                external_info = {'status': resp.status_code, 'data': resp_json}

                try:
                    if isinstance(resp_json, dict):
                        codes = resp_json.get('codes')
                        if codes and isinstance(codes, list) and len(codes) > 0:
                            first = codes[0]
                            created_at_str = first.get('created_at')
                            email = first.get('email')
                            external_info['cliente'] = email
                            external_info['sn'] = first.get('sn')

                            if created_at_str:
                                try:
                                    created_dt = parsedate_to_datetime(created_at_str)
                                    try:
                                        venc_dt = created_dt.replace(year=created_dt.year + 2)
                                    except ValueError:
                                        venc_dt = created_dt.replace(month=2, day=28, year=created_dt.year + 2)
                                    external_info['created_at'] = created_dt.isoformat()
                                    external_info['vencimento'] = venc_dt.date().isoformat()
                                except Exception:
                                    pass
                except Exception:
                    pass

                context['external_search'] = external_info

            except requests.RequestException as exc:
                context['external_search'] = {'error': str(exc)}

            found_external = False
            es = context.get('external_search')
            if isinstance(es, dict):
                data = es.get('data')
                if isinstance(data, dict):
                    codes = data.get('codes')
                    if codes and isinstance(codes, list) and len(codes) > 0:
                        found_external = True

            if found_external:
                context['mensagem'] = 'Dados captados externamente, necessário atualização'

                data = es.get('data') if isinstance(es, dict) else None
                first = None
                if isinstance(data, dict):
                    codes = data.get('codes')
                    if codes and isinstance(codes, list) and len(codes) > 0:
                        first = codes[0]

                cliente_dict = {
                    'nome': es.get('cliente') or (first.get('email') if first else '') or '',
                    'cnpj': '',
                    'tel': '',
                    'equipamento': ('{} - {}'.format(first.get('city', ''), first.get('country', '')).strip(' -') if first else ''),
                    'vencimento': es.get('vencimento') or (first.get('created_at') if first else None),
                }
                context['cliente'] = cliente_dict

                status_val = 'indefinido'
                venc_str = es.get('vencimento') if isinstance(es, dict) else None

                if venc_str:
                    try:
                        venc_date = parse_date(venc_str)
                        if isinstance(venc_date, datetime):
                            venc_date = venc_date.date()
                        dias = (venc_date - date.today()).days
                        if dias > 30:
                            status_val = 'direito'
                        elif dias < 1:
                            status_val = 'vencido'
                        else:
                            status_val = 'vencendo'
                    except Exception:
                        status_val = 'indefinido'
                else:
                    created_at_str = first.get('created_at') if first else None
                    if created_at_str:
                        try:
                            created_dt = parsedate_to_datetime(created_at_str)
                            try:
                                venc_dt = created_dt.replace(year=created_dt.year + 2)
                            except ValueError:
                                venc_dt = created_dt.replace(month=2, day=28, year=created_dt.year + 2)
                            dias = (venc_dt.date() - date.today()).days
                            if dias > 30:
                                status_val = 'direito'
                            elif dias < 1:
                                status_val = 'vencido'
                            else:
                                status_val = 'vencendo'
                            if not cliente_dict.get('vencimento'):
                                cliente_dict['vencimento'] = venc_dt.date().isoformat()
                        except Exception:
                            status_val = 'indefinido'

                context['status'] = status_val
                if status_val == 'direito':
                    context['status_message'] = "SUPORTE LIBERADO - Atualizar dados e atender normalmente"
                elif status_val == 'vencido':
                    context['status_message'] = "SUPORTE VENCIDO - Não fazer atendimento"
                elif status_val == 'vencendo':
                    context['status_message'] = "SUPORTE A VENCER - Atualizar dados e atender normalmente"

            else:
                context['status_message'] = 'SEM DADOS'
                context['mensagem'] = 'Passar para o comercial atualizar o cadastro.'
                SerialSearchLog.objects.create(
                    user=user,
                    searched_serial=serial_input,
                    resolved_serial=""
                )

            return render(request, 'situacao/index.html', context)

        # =======================
        # ENCONTROU NO BANCO
        # =======================
        if clientes.count() > 1:
            lista_clientes = []
            for cliente in clientes:
                lista_clientes.append({
                    'cliente': cliente,
                    'status': cliente.status,
                    'vencimento_dias': cliente._vencimento_dias,
                    'status_message': cliente.status_message,
                    'serial_principal': cliente.serial,
                    'serial_secundario': cliente.serial_sec,
                })
            context['clientes_duplicados'] = lista_clientes
            context['mensagem'] = 'Encontradas múltiplas ocorrências para esse serial. Verifique os dados abaixo:'
            return render(request, 'situacao/index.html', context)

        cliente = clientes.first()
        SerialSearchLog.objects.create(
            user=user,
            searched_serial=serial_input,
            resolved_serial=cliente.serial or ""
        )

        ui_confirmar_contactado = False

        if cliente.status == "vencido" and not getattr(cliente, "contactado", False):
            # backend marca automaticamente
            cliente.contactado = True
            cliente.updated_by = request.user if request.user.is_authenticated else None
            cliente.save(update_fields=["contactado", "updated_by", "updated_at"])

            admin_log(
                user=request.user if request.user.is_authenticated else None,
                obj=cliente,
                action_flag=CHANGE,
                message=(
                    f"[serial={cliente.serial}] Primeira consulta com suporte vencido: "
                    f"(digitado='{serial_input}')"
                )
            )

            # ✅ só nessa resposta, permite o “botão visual”
            ui_confirmar_contactado = True

            # ✅ força a mensagem desta resposta (mesmo já estando contactado no DB)
            context["status_message"] = "Suporte vencido - último atendimento"

        buscou_pelo_sec = bool(cliente.serial_sec) and serial_input.casefold() == cliente.serial_sec.casefold()
        if buscou_pelo_sec and cliente.serial:
            context['serial_digitado'] = cliente.serial

        context['serial_buscado'] = serial_input
        context['serial_principal'] = cliente.serial
        context['serial_secundario'] = cliente.serial_sec
        context['buscou_pelo_secundario'] = buscou_pelo_sec and (serial_input.casefold() != (cliente.serial or '').casefold())

        context["cliente"] = cliente
        context["status"] = cliente.status
        context["mensagem"] = cliente.message_effective

        # se não for a primeira vez, usa o status_message normal
        context["status_message"] = context.get("status_message") or cliente.status_message

        # ✅ manda a flag pra UI
        context["ui_confirmar_contactado"] = ui_confirmar_contactado

        return render(request, "situacao/index.html", context)

    return render(request, "situacao/index.html", context)


def _anos_por_equipamento(equipamento: str) -> int:
    if not equipamento:
        return 2
    return 1 if equipamento.lower().find('reader') != -1 else 2


ALLOWED_FIELDS = {"nome", "cnpj", "tel", "vencimento", "serial_sec", "contactado"}


def _digits_only(s: str) -> str:
    return ''.join(ch for ch in s if ch.isdigit())


def _parse_excel_date(value, workbook):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, epoch=workbook.epoch).date()
        except Exception:
            return None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        parsed = parse_date(cleaned)
        if parsed:
            return parsed
        try:
            return datetime.strptime(cleaned, "%d/%m/%Y").date()
        except ValueError:
            return None
    return None


@require_POST
@login_required(login_url=settings.URL_LOGIN)
@require_system_access("situacao")
def cadastrar_serial(request):
    data_referencia = timezone.now()

    serial = (request.POST.get('serial') or '').strip()
    serial_sec = (request.POST.get('serial_sec') or '').strip()

    nome = (request.POST.get('nome') or '').strip()
    cnpj = (request.POST.get('cnpj') or '').strip()
    tel = (request.POST.get('tel') or '').strip()
    equipamento = (request.POST.get('equipamento') or '').strip()
    data_input = (request.POST.get('data') or '').strip()

    anos_input = (request.POST.get('anos_para_vencimento') or '').strip()
    venc_input = (request.POST.get('vencimento') or '').strip()

    field_errors = {}
    if not serial:
        field_errors['serial'] = 'Serial é obrigatório.'

    if serial and serial_sec and serial.casefold() == serial_sec.casefold():
        field_errors['serial_sec'] = 'Serial secundário não pode ser igual ao serial principal.'

    anos_para_vencimento = None
    vencimento_data = None

    if check_user_full_permission(request.user):
        if anos_input:
            try:
                anos_para_vencimento = int(anos_input)
                if anos_para_vencimento < 0:
                    field_errors['anos_para_vencimento'] = 'Deve ser um inteiro zero ou positivo.'
            except ValueError:
                field_errors['anos_para_vencimento'] = 'Informe um número inteiro.'
        else:
            anos_para_vencimento = 2

        if venc_input:
            d = parse_date(venc_input)
            if not d:
                field_errors['vencimento'] = 'Data inválida (use YYYY-MM-DD).'
            else:
                vencimento_data = d

    if field_errors:
        return JsonResponse(
            {"ok": False, "message": "Corrija os campos destacados.", "field_errors": field_errors},
            status=400,
        )

    try:
        if Cliente.objects.filter(Q(serial__iexact=serial) | Q(serial_sec__iexact=serial)).exists():
            return JsonResponse(
                {"ok": False, "message": "Serial já em uso.", "field_errors": {"serial": "Serial já cadastrado (principal ou secundário)."}},
                status=409,
            )

        if serial_sec and Cliente.objects.filter(Q(serial__iexact=serial_sec) | Q(serial_sec__iexact=serial_sec)).exists():
            return JsonResponse(
                {"ok": False, "message": "Serial secundário já em uso.", "field_errors": {"serial_sec": "Serial secundário já cadastrado (principal ou secundário)."}},
                status=409,
            )

        if not check_user_full_permission(request.user):
            anos_para_vencimento = _anos_por_equipamento(equipamento)

        data_lanc = None
        if data_input:
            d = parse_date(data_input)
            if not d:
                return JsonResponse(
                    {"ok": False, "message": "Data inválida (use YYYY-MM-DD).", "field_errors": {"data": "Data inválida."}},
                    status=400,
                )
            data_lanc = d

        cliente = Cliente.objects.create(
            data=data_lanc or timezone.now().date(),
            anos_para_vencimento=int(anos_para_vencimento),
            serial=serial,
            serial_sec=serial_sec or "",
            nome=nome,
            cnpj=_digits_only(cnpj) or cnpj,
            tel=tel,
            equipamento=equipamento or "N/D",
            vencimento=vencimento_data if check_user_full_permission(request.user) else None,
            created_by=request.user if request.user.is_authenticated else None,
            contactado=False,
        )

        admin_log(
            user=request.user,
            obj=cliente,
            action_flag=ADDITION,
            message=(
                f"Criou cliente via cadastro (form/painel). "
                f"ID={cliente.id} | serial='{cliente.serial}'"
                f"{' | serial_sec=' + repr(cliente.serial_sec) if cliente.serial_sec else ''} "
                f"| nome='{(cliente.nome or '').strip()}'"
            )
        )

        return JsonResponse(
            {
                "ok": True,
                "message": "Cadastro realizado com sucesso!",
                "data": {
                    "id": cliente.id,
                    "serial": cliente.serial,
                    "serial_sec": cliente.serial_sec,
                    "nome": cliente.nome,
                    "cnpj": cliente.cnpj,
                    "tel": cliente.tel,
                    "equipamento": cliente.equipamento,
                    "anos_para_vencimento": cliente.anos_para_vencimento,
                    "vencimento": cliente.vencimento.isoformat() if cliente.vencimento else None,
                    "contactado": cliente.contactado,
                    "timestamp": data_referencia.isoformat(),
                },
            },
            status=201,
        )

    except IntegrityError:
        return JsonResponse(
            {"ok": False, "message": "Não foi possível cadastrar (restrição de unicidade)."},
            status=409,
        )
    except ValueError as e:
        return JsonResponse(
            {"ok": False, "message": f"Erro de validação: {e}"},
            status=400,
        )
    except Exception:
        return JsonResponse(
            {"ok": False, "message": "Erro interno ao cadastrar."},
            status=500,
        )


@require_GET
def api_buscar_cliente(request):
    serial = (request.GET.get('serial') or '').strip()
    if not serial:
        return JsonResponse({"ok": False, "message": "Informe o serial."}, status=400)

    cliente = _first_by_any_id(serial)
    if not cliente:
        return JsonResponse({"ok": False, "message": "Serial não encontrado."}, status=404)

    data = {
        "id": cliente.id,
        "serial": cliente.serial,
        "serial_sec": cliente.serial_sec,
        "nome": cliente.nome,
        "cnpj": cliente.cnpj,
        "tel": cliente.tel,
        "vencimento": cliente.vencimento.isoformat() if cliente.vencimento else None,
        "contactado": getattr(cliente, "contactado", False),
    }
    return JsonResponse({"ok": True, "data": data}, status=200)


@require_POST
def api_atualizar_cliente(request):
    serial = (request.POST.get('serial') or '').strip()
    field = (request.POST.get('field') or '').strip()
    value = (request.POST.get('value') or '')

    if not serial:
        return JsonResponse({"ok": False, "message": "Serial é obrigatório."}, status=400)
    if field not in ALLOWED_FIELDS:
        return JsonResponse({"ok": False, "message": "Campo não permitido para atualização."}, status=400)

    cliente = _first_by_any_id(serial)
    if not cliente:
        return JsonResponse({"ok": False, "message": "Serial não encontrado."}, status=404)

    if field == "cnpj":
        value = ''.join(ch for ch in value if ch.isdigit()) or value

    if field == "contactado":
        raw = str(value or "").strip().lower()
        true_set = {"1", "true", "t", "sim", "yes", "y", "on"}
        false_set = {"0", "false", "f", "nao", "não", "no", "n", "off", ""}

        if raw in true_set:
            value_bool = True
        elif raw in false_set:
            value_bool = False
        else:
            return JsonResponse(
                {"ok": False, "message": "Valor inválido para contactado (use true/false)."},
                status=400
            )

        cliente.contactado = value_bool
        cliente.updated_by = request.user if request.user.is_authenticated else None
        cliente.save(update_fields=["contactado", "updated_by", "updated_at"])

        admin_log(
            user=request.user,
            obj=cliente,
            action_flag=CHANGE,
            message=f"Atualizou contactado para '{value_bool}' via API"
        )

        return JsonResponse({"ok": True, "message": "Contactado atualizado.", "data": {"contactado": value_bool}})

    if field == "serial_sec":
        new_val = (value or "").strip()

        if not new_val:
            cliente.serial_sec = ""
            cliente.updated_by = request.user if request.user.is_authenticated else None
            cliente.save(update_fields=["serial_sec", "updated_by", "updated_at"])
            admin_log(
                user=request.user,
                obj=cliente,
                action_flag=CHANGE,
                message="Limpou serial_sec via API"
            )
            return JsonResponse({"ok": True, "message": "Serial secundário removido.", "data": {"serial_sec": ""}})

        if cliente.serial and new_val.casefold() == cliente.serial.casefold():
            return JsonResponse({"ok": False, "message": "Serial secundário não pode ser igual ao serial principal."}, status=400)

        exists = Cliente.objects.filter(
            Q(serial__iexact=new_val) | Q(serial_sec__iexact=new_val)
        ).exclude(pk=cliente.pk).exists()
        if exists:
            return JsonResponse({"ok": False, "message": "Este serial secundário já está em uso."}, status=409)

        cliente.serial_sec = new_val
        cliente.updated_by = request.user if request.user.is_authenticated else None
        cliente.save(update_fields=["serial_sec", "updated_by", "updated_at"])
        admin_log(
            user=request.user,
            obj=cliente,
            action_flag=CHANGE,
            message=f"Atualizou serial_sec para '{new_val}' via API"
        )
        return JsonResponse({"ok": True, "message": "Serial secundário atualizado.", "data": {"serial_sec": new_val}})

    if field == "vencimento":
        if not value:
            cliente.vencimento = None
            cliente.updated_by = request.user if request.user.is_authenticated else None
            cliente.save(update_fields=["vencimento", "updated_by", "updated_at"])
            return JsonResponse({"ok": True, "message": "Vencimento removido."})

        data_v = parse_date(value)
        if not data_v:
            return JsonResponse({"ok": False, "message": "Data inválida. Use formato YYYY-MM-DD."}, status=400)

        cliente.vencimento = data_v
        cliente.updated_by = request.user if request.user.is_authenticated else None
        cliente.save(update_fields=["vencimento", "updated_by", "updated_at"])
        return JsonResponse({"ok": True, "message": "Vencimento atualizado.", "data": {"vencimento": data_v.isoformat()}})

    setattr(cliente, field, value)
    cliente.updated_by = request.user if request.user.is_authenticated else None
    cliente.save(update_fields=[field, "updated_by", "updated_at"])

    admin_log(
        user=request.user,
        obj=cliente,
        action_flag=CHANGE,
        message=f"Atualizou campo '{field}' para '{value}' via API"
    )

    return JsonResponse({"ok": True, "message": f"{field.capitalize()} atualizado.", "data": {field: value}})


@require_POST
@login_required(login_url=settings.URL_LOGIN)
@require_system_access("situacao")
def importar_excel(request):
    planilha = request.FILES.get('arquivo_excel')
    if not planilha:
        return JsonResponse({"ok": False, "message": "Envie um arquivo Excel."}, status=400)

    try:
        workbook = load_workbook(planilha, data_only=True)
    except Exception:
        return JsonResponse({"ok": False, "message": "Não foi possível ler o arquivo enviado."}, status=400)

    sheet = workbook.active
    header_cells = sheet[1]

    expected = {
        "nome cliente": "nome",
        "nome item": "equipamento",
        "serial": "serial",
        "cnpj/cpf": "cnpj",
        "contato": "tel",
        "numero emissão nf": "data",
    }

    header_map = {}
    for idx, cell in enumerate(header_cells):
        header = str(cell.value).strip().lower() if cell.value else ""
        if header in expected:
            header_map[expected[header]] = idx

    if "serial" not in header_map:
        return JsonResponse(
            {"ok": False, "message": "A coluna 'serial' é obrigatória na planilha."},
            status=400,
        )

    created = 0
    duplicates = 0
    errors = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        row_data = {}
        for field, col_idx in header_map.items():
            value = row[col_idx].value if col_idx < len(row) else None
            row_data[field] = value

        if not any(row_data.values()):
            continue

        serial = (row_data.get("serial") or "").strip()
        if not serial:
            errors.append({"row": row_index, "message": "Serial ausente."})
            continue

        if Cliente.objects.filter(Q(serial__iexact=serial) | Q(serial_sec__iexact=serial)).exists():
            duplicates += 1
            continue

        nome = (row_data.get("nome") or "").strip()
        equipamento = (row_data.get("equipamento") or "").strip() or "N/D"
        cnpj = _digits_only((row_data.get("cnpj") or "").strip())
        tel = (row_data.get("tel") or "").strip()

        data_valor = _parse_excel_date(row_data.get("data"), workbook)
        data_lanc = data_valor or timezone.localdate()

        try:
            Cliente.objects.create(
                data=data_lanc,
                anos_para_vencimento=_anos_por_equipamento(equipamento),
                serial=serial,
                nome=nome,
                cnpj=cnpj or None,
                tel=tel or None,
                equipamento=equipamento,
                created_by=request.user if request.user.is_authenticated else None,
                contactado=False,
            )
            created += 1
        except IntegrityError:
            duplicates += 1
        except Exception as exc:
            errors.append({"row": row_index, "message": str(exc)})

    message = "Importação concluída."
    if created or duplicates or errors:
        message = (
            f"Importação concluída. Criados: {created}. Duplicados ignorados: {duplicates}. "
            f"Erros: {len(errors)}."
        )

    status_code = 200 if not errors else 207
    return JsonResponse(
        {"ok": True, "message": message, "data": {"created": created, "duplicates": duplicates, "errors": errors}},
        status=status_code,
    )


@require_GET
def equipamentos_suggest(request):
    q = (request.GET.get('q') or '').strip()
    try:
        limit = int(request.GET.get('limit') or 15)
    except ValueError:
        limit = 15
    limit = max(1, min(limit, 50))

    def _norm(s: str) -> str:
        if not s:
            return ''
        s = unicodedata.normalize('NFKD', s)
        s = ''.join(ch for ch in s if not unicodedata.combining(ch))
        s = ''.join(ch for ch in s if ch.isalnum())
        return s.casefold()

    norm_q = _norm(q)

    base = list(Device.objects.order_by('name').values_list('name', flat=True))
    if norm_q:
        filtered = [name for name in base if norm_q in _norm(name)]
    else:
        filtered = base

    return JsonResponse({"results": filtered[:limit]})


@require_GET
def odoo_sync(request):
    try:
        raw_limit = request.GET.get('limit')
        if not raw_limit or str(raw_limit).lower() in ('all', 'ilimitado', 'none', 'null'):
            eff_limit = None
        else:
            try:
                eff_limit = int(raw_limit)
            except ValueError:
                eff_limit = None

        used_url = getattr(settings, 'ODOO_URL', None) or os.getenv('ODOO_URL')
        used_db = getattr(settings, 'ODOO_DB', None) or os.getenv('ODOO_DB')

        stats = sync_odoo_to_clientes(max_rows=eff_limit)
        return JsonResponse({"ok": True, "message": "Sync concluído", "data": {**stats, "using_url": used_url, "using_db": used_db}})

    except (requests.exceptions.RequestException, ConnectionError) as e:
        used_url = getattr(settings, 'ODOO_URL', None) or os.getenv('ODOO_URL')
        used_db = getattr(settings, 'ODOO_DB', None) or os.getenv('ODOO_DB')
        return JsonResponse({"ok": False, "message": f"Falha ao conectar ao Odoo: {e}", "using_url": used_url, "using_db": used_db}, status=502)
    except Exception as e:
        return JsonResponse({"ok": False, "message": str(e)}, status=500)


def _parse_iso_datetime(value: str):
    if not value:
        return None
    dt = parse_datetime(value)
    if not dt:
        return None
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone=timezone.utc)
    return dt


import secrets

@csrf_exempt
@require_POST
def webhook_equipment_status(request):
    expected_token = getattr(settings, "SITUACAO_WEBHOOK_TOKEN", None)
    if expected_token:
        got = request.headers.get("X-Webhook-Token", "")
        if not secrets.compare_digest(got, expected_token):
            return JsonResponse(
                {"success": False, "error": {"code": "UNAUTHORIZED", "message": "Invalid webhook token"}},
                status=401
            )

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return JsonResponse(
            {"success": False, "error": {"code": "BAD_JSON", "message": "Invalid JSON"}},
            status=400
        )

    serial = (payload.get("serialNumber") or "").strip()
    status_str = (payload.get("status") or "").strip().upper()
    blocked_by = (payload.get("blockedBy") or "").strip()
    reason = payload.get("reason")
    event = payload.get("event")
    ts = payload.get("timestamp")

    if not serial:
        return JsonResponse(
            {"success": False, "error": {"code": "MISSING_SERIAL", "message": "serialNumber is required"}},
            status=400
        )

    cliente = _first_by_any_id(serial)
    if not cliente:
        return JsonResponse(
            {"success": False, "error": {"code": "NOT_FOUND", "message": f"Serial {serial} not found"}},
            status=404
        )

    dt = _parse_iso_datetime(ts) or timezone.now()
    event_date = timezone.localtime(dt).date()

    updated_user = None
    if blocked_by:
        updated_user = User.objects.filter(username__iexact=blocked_by).first()

    try:
        cliente.equipment_status = status_str
        cliente.equipment_reason = reason
        cliente.equipment_blocked_by = blocked_by
        cliente.equipment_last_update = dt
        cliente.last_webhook_event = event
    except Exception:
        pass

    # dt = _parse_iso_datetime(ts) or timezone.now()
    # event_date = timezone.localtime(dt).date()

    if status_str == "BLOCKED":
        # usa a data do evento (timestamp do webhook)
        cliente.data = event_date

        # opcional: não mexer em vencimento (pra não bagunçar suporte normal)
        # mas se sua tela depende de vencimento, pode setar igual event_date:
        # cliente.vencimento = event_date

        cliente.status_message_custom = None  # deixa o model controlar
        cliente.contactado = False

    elif status_str == "UNBLOCKED":
        cliente.data = event_date
        cliente.anos_para_vencimento = 2
        cliente.vencimento = cliente.data + relativedelta(years=2)

        # ✅ limpa mensagem custom ao desbloquear
        cliente.status_message_custom = None
        cliente.contactado = False
        # aqui você decide se mantém contactado ou não
    cliente.updated_by = updated_user

    fields = [
        "vencimento",
        "data",
        "anos_para_vencimento",
        "updated_by",
        "updated_at",
        "contactado",
    ]

    for f in ["equipment_status", "equipment_reason", "equipment_blocked_by", "equipment_last_update", "last_webhook_event", "status_message_custom"]:
        if hasattr(cliente, f):
            fields.append(f)

    cliente.save(update_fields=list(dict.fromkeys(fields)))

    admin_log(
        user=updated_user,
        obj=cliente,
        action_flag=CHANGE,
        message=f"[serial={cliente.serial}] Atualizou cliente via webhook. status={status_str} blocked_by='{blocked_by}'"
    )

    return JsonResponse({
        "success": True,
        "data": {
            "serialNumber": cliente.serial,
            "serial_sec": getattr(cliente, "serial_sec", ""),
            "status": status_str,
            "vencimento": cliente.vencimento.isoformat() if cliente.vencimento else None,
            "data": cliente.data.isoformat() if cliente.data else None,
            "computed_status": getattr(cliente, "status", None),
            "status_message": getattr(cliente, "status_message", None),
            "contactado": getattr(cliente, "contactado", False),
            "updated_by": cliente.updated_by.username if cliente.updated_by else payload.get("blockedBy"),
        }
    })

def is_gestor_ti(user):
	profile = getattr(user, 'profile', None)
	return bool(
		getattr(profile, 'role', None) == 'gestor'
		and getattr(profile, 'setor', None) == 'ti'
	)

@login_required(login_url=settings.URL_LOGIN)
@require_system_access("situacao")
def index(request):
    profile = getattr(request.user, 'profile', None)
    user_role = getattr(profile, 'role', None)
    is_master = (
        user_role in ["dono", "diretor", "gestor"] or is_gestor_ti(request.user)
    )
    return render(request, 'situacao/index.html', {'clientes': None, 'is_master':is_master})

@require_POST
@login_required(login_url=settings.URL_LOGIN)
@require_system_access("situacao")
def marcar_contactado(request):
    serial = (request.POST.get("serial") or "").strip()
    if not serial:
        return JsonResponse({"ok": False, "message": "Informe o serial."}, status=400)

    cliente = _first_by_any_id(serial)
    if not cliente:
        return JsonResponse({"ok": False, "message": "Serial não encontrado."}, status=404)

    # só permite marcar se estiver vencido (recomendado)
    if cliente.status != "vencido":
        return JsonResponse(
            {"ok": False, "message": "Só é possível marcar 'contactado' quando o suporte estiver VENCIDO."},
            status=400
        )

    cliente.contactado = True
    cliente.updated_by = request.user
    cliente.save(update_fields=["contactado", "updated_by", "updated_at"])

    admin_log(
        user=request.user,
        obj=cliente,
        action_flag=CHANGE,
        message=f"[serial={cliente.serial}] Marcou contactado=True via checkbox (suporte vencido)."
    )

    return JsonResponse({"ok": True, "message": "Marcado como contactado.", "data": {"contactado": True}})
